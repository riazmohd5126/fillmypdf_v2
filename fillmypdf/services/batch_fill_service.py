"""
Batch Fill Service
==================
Handle bulk PDF filling from JSON arrays and CSV files
"""

import csv
import io
import zipfile
import json
from pathlib import Path
from typing import Callable, Dict, Any, Optional, List
from datetime import datetime

from .vision_service import VisionService
from .pdf_service import PDFService
from .profile_service import ProfileService
from .input_adapter import InputAdapter
from ..config import settings


class BatchFillService:
    """Service for batch PDF filling operations"""
    
    def __init__(self):
        self.pdf_service = PDFService()
        self.profile_service = ProfileService()
        self.input_adapter = InputAdapter()

    def analyze_template_fields(self, template_pdf_path: Path) -> Dict[str, Any]:
        """
        Convert template to fillable (same as batch) then list detected fields +
        inferred labels — no AI key required.

        Enables clients to validate that commonforms/detection sees the PDF
        before paying for LLM mappings.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fillable_path = settings.UPLOAD_DIR / f"{timestamp}_inspect_fillable.pdf"

        ok = self.pdf_service.convert_to_fillable(
            input_path=str(template_pdf_path),
            output_path=str(fillable_path),
        )
        if not ok:
            return {
                "success": False,
                "fields_detected": 0,
                "fields": [],
                "message": "Failed to convert template to fillable PDF",
            }

        # No API calls — placeholder constructor args are ignored by inspect
        vision = VisionService(api_key="-", base_url="https://example.invalid", model="none")
        try:
            data = vision.inspect_fillable_form(str(fillable_path))
        finally:
            fillable_path.unlink(missing_ok=True)

        return {
            "success": True,
            "fields_detected": data["fields_detected"],
            "fields": data["fields"],
            "message": None,
        }
    
    def parse_csv(self, csv_content: bytes) -> List[Dict[str, str]]:
        """
        Parse CSV file into list of records
        
        Args:
            csv_content: Raw CSV file bytes
            
        Returns:
            List of dictionaries, one per row
        """
        try:
            csv_text = csv_content.decode('utf-8')
        except UnicodeDecodeError:
            # Try common encodings
            for encoding in ['utf-8-sig', 'latin-1', 'cp1252']:
                try:
                    csv_text = csv_content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ValueError("Could not decode CSV file")
        
        csv_reader = csv.DictReader(io.StringIO(csv_text))
        records = list(csv_reader)
        
        if not records:
            raise ValueError("CSV file is empty or has no data rows")
        
        return records
    
    def process_batch_json(
        self,
        template_pdf_path: Path,
        user_data_array: List[Dict[str, str]],
        ai_api_key: str,
        ai_base_url: str,
        ai_model: str,
        batch_id: str,
        dpi: int = 200,
        profile_id: Optional[str] = None,
        profile_ids: Optional[List[str]] = None,
        on_record_done: Optional[Callable[[int, int, int], None]] = None,
    ) -> Dict[str, Any]:
        """
        Process batch fill from JSON array.
        Pass profile_ids (list) to merge multiple profiles; falls back to profile_id.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Resolve profiles — multi takes precedence
        base_profile_data = {}
        ids = profile_ids or ([profile_id] if profile_id else [])
        if len(ids) > 1:
            try:
                base_profile_data = self.profile_service.use_profiles(ids)
                print(f"✅ Using profiles: {ids}")
            except Exception as e:
                print(f"⚠️  Profile merge failed: {e}")
        elif ids:
            try:
                base_profile_data = self.profile_service.use_profile(ids[0])
                print(f"✅ Using profile: {ids[0]}")
            except ValueError as e:
                print(f"⚠️  Profile not found: {e}")
        
        # Step 1: Convert template to fillable (once)
        fillable_path = settings.UPLOAD_DIR / f"{timestamp}_{batch_id}_fillable.pdf"
        
        convert_success = self.pdf_service.convert_to_fillable(
            input_path=template_pdf_path,
            output_path=fillable_path
        )
        
        if not convert_success:
            raise ValueError("Failed to convert PDF template to fillable form")
        
        # Step 2: Create batch directory
        batch_dir = settings.OUTPUT_DIR / f"{timestamp}_{batch_id}_batch"
        batch_dir.mkdir(exist_ok=True)
        
        # Step 3: Initialize Vision service
        vision = VisionService(
            api_key=ai_api_key,
            base_url=ai_base_url,
            model=ai_model
        )
        
        # Step 4: Process each record
        results = []
        successful = 0
        failed = 0
        
        print(f"\n{'='*60}")
        print(f"BATCH JOB {batch_id} - {len(user_data_array)} RECORDS")
        print(f"{'='*60}\n")
        
        for idx, user_data in enumerate(user_data_array, 1):
            try:
                # Merge profile data with record data (record overrides profile)
                # Then run through InputAdapter for normalisation + alias emission
                ai_input = self.input_adapter.to_ai_input(user_data, base_profile_data)
                merged_data = ai_input  # canonical bundle {"flat": ..., "structured": ...}

                # Generate filename from flat data
                filename = self._generate_filename(ai_input["flat"], idx)
                output_path = batch_dir / filename

                # Fill PDF
                result = vision.autofill_pipeline(
                    fillable_pdf_path=str(fillable_path),
                    output_path=str(output_path),
                    user_data=merged_data,
                    dpi=dpi
                )
                
                avg_conf = result.get("avg_confidence")
                cache_hit = result.get("cache_hit", False)
                if result["success"]:
                    successful += 1
                    conf_str = f"  conf={avg_conf:.2f}" if avg_conf is not None else ""
                    cache_str = "  [cache]" if cache_hit else ""
                    print(f"  [{idx}/{len(user_data_array)}] \u2705 {filename}{conf_str}{cache_str}")
                else:
                    failed += 1
                    print(f"  [{idx}/{len(user_data_array)}] \u274c {filename} - {result.get('error')}")

                results.append({
                    "index": idx,
                    "filename": filename,
                    "success": result["success"],
                    "fields_detected": result.get("fields_detected", 0),
                    "fields_filled": result.get("fields_filled", 0),
                    "avg_confidence": avg_conf,
                    "cache_hit": cache_hit,
                    "field_labels": result.get("field_labels", {}),
                    "mappings": result.get("mappings", {}),
                    "confidence": result.get("confidence", {}),
                    "error": result.get("error"),
                })
            
            except Exception as e:
                failed += 1
                print(f"  [{idx}/{len(user_data_array)}] ❌ Error: {e}")
                results.append({
                    "index": idx,
                    "success": False,
                    "error": str(e),
                })
            
            if on_record_done is not None:
                on_record_done(idx, successful, failed)

        # Step 5: Create ZIP archive
        zip_filename = f"batch_{timestamp}_{batch_id}.zip"
        zip_path = settings.OUTPUT_DIR / zip_filename
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add successful PDFs
            for result in results:
                if result.get("success") and result.get("filename"):
                    pdf_path = batch_dir / result["filename"]
                    if pdf_path.exists():
                        zipf.write(pdf_path, result["filename"])
            
            # Add batch report
            report_data = {
                "job_id": batch_id,
                "timestamp": timestamp,
                "total_records": len(user_data_array),
                "successful": successful,
                "failed": failed,
                "profile_used": profile_id,
                "results": results,
            }
            zipf.writestr("batch_report.json", json.dumps(report_data, indent=2))
        
        # Step 6: Cleanup
        fillable_path.unlink(missing_ok=True)
        # Keep batch_dir for now (cleanup in background task)
        
        print(f"\n{'='*60}")
        print(f"BATCH COMPLETE - {successful}/{len(user_data_array)} successful")
        print(f"{'='*60}\n")
        
        return {
            "batch_id": batch_id,
            "total_records": len(user_data_array),
            "successful": successful,
            "failed": failed,
            "success_rate": round(successful / len(user_data_array) * 100, 1) if user_data_array else 0,
            "zip_file": str(zip_path),
            "zip_filename": zip_filename,
            "download_url": f"/api/v1/batch/download/{zip_filename}",
            "results": results,
            "batch_dir": str(batch_dir),  # For cleanup
        }
    
    def process_csv_batch(
        self,
        template_pdf_path: Path,
        csv_content: bytes,
        csv_filename: str,
        ai_api_key: str,
        ai_base_url: str,
        ai_model: str,
        batch_id: str,
        dpi: int = 200,
        profile_id: Optional[str] = None,
        profile_ids: Optional[List[str]] = None,
        on_record_done: Optional[Callable[[int, int, int], None]] = None,
    ) -> Dict[str, Any]:
        records = self.parse_csv(csv_content)
        if len(records) > 500:
            raise ValueError("Maximum 500 rows per CSV batch")
        print(f"📊 CSV parsed: {len(records)} records")
        print(f"   Columns: {list(records[0].keys())}")
        return self.process_batch_json(
            template_pdf_path=template_pdf_path,
            user_data_array=records,
            ai_api_key=ai_api_key,
            ai_base_url=ai_base_url,
            ai_model=ai_model,
            batch_id=batch_id,
            dpi=dpi,
            profile_id=profile_id,
            profile_ids=profile_ids,
            on_record_done=on_record_done,
        )

    def parse_xlsx(self, xlsx_content: bytes) -> List[Dict[str, str]]:
        """
        Parse the first worksheet of an .xlsx into one dict per row using
        header row keys (same semantics as CSV batch).
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError(
                "openpyxl is not installed — add openpyxl to requirements for Excel batches"
            )
        bio = io.BytesIO(xlsx_content)
        wb = load_workbook(bio, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not rows:
            raise ValueError("Excel workbook is empty")
        hdr = rows[0]
        headers = [
            str(h).strip() if h is not None and str(h).strip() else f"column_{idx}"
            for idx, h in enumerate(hdr)
        ]
        records: List[Dict[str, str]] = []
        for row in rows[1:]:
            rec: Dict[str, str] = {}
            nonempty = False
            row = row or ()
            for idx, head in enumerate(headers):
                raw = row[idx] if idx < len(row) else None
                if isinstance(raw, bool):
                    txt = "TRUE" if raw else "FALSE"
                elif raw is None:
                    txt = ""
                elif isinstance(raw, int):
                    txt = str(raw)
                elif isinstance(raw, float) and raw == int(raw):
                    txt = str(int(raw))
                elif isinstance(raw, float):
                    txt = str(raw).strip()
                else:
                    txt = str(raw).strip()
                rec[head] = txt
                if txt:
                    nonempty = True
            if nonempty:
                records.append(rec)
        if not records:
            raise ValueError("Excel file has no populated data rows after the header")
        return records

    def process_xlsx_batch(
        self,
        template_pdf_path: Path,
        xlsx_content: bytes,
        xlsx_filename: str,
        ai_api_key: str,
        ai_base_url: str,
        ai_model: str,
        batch_id: str,
        dpi: int = 200,
        profile_id: Optional[str] = None,
        profile_ids: Optional[List[str]] = None,
        on_record_done: Optional[Callable[[int, int, int], None]] = None,
    ) -> Dict[str, Any]:
        records = self.parse_xlsx(xlsx_content)
        if len(records) > 500:
            raise ValueError("Maximum 500 rows per Excel batch")
        print(f"📊 Excel ({xlsx_filename}) parsed: {len(records)} records")
        print(f"   Columns: {list(records[0].keys())}")
        return self.process_batch_json(
            template_pdf_path=template_pdf_path,
            user_data_array=records,
            ai_api_key=ai_api_key,
            ai_base_url=ai_base_url,
            ai_model=ai_model,
            batch_id=batch_id,
            dpi=dpi,
            profile_id=profile_id,
            profile_ids=profile_ids,
            on_record_done=on_record_done,
        )

    def _generate_filename(self, data: Dict[str, str], index: int) -> str:
        """Generate filename from user data"""
        name_parts = []
        
        # Try common name fields
        for key in ['first_name', 'last_name', 'name', 'full_name', 'FirstName', 'LastName']:
            if key in data and data[key]:
                name_parts.append(str(data[key]).replace(' ', '_'))
        
        if name_parts:
            filename = '_'.join(name_parts)
        else:
            filename = f"record_{index}"
        
        # Sanitize filename
        filename = "".join(c for c in filename if c.isalnum() or c in ('_', '-'))
        
        return f"{filename}.pdf"


# Example usage
if __name__ == "__main__":
    print("BatchFillService - use via API endpoints")
