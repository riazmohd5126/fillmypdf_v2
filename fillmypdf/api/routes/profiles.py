"""
Profile API Routes
==================
CRUD endpoints for user profiles (scoped to the calling API key).
"""

import csv
import io
from typing import Any, Callable, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile

from ...models import Profile, ProfileCreate, ProfileUpdate
from ...services.profile_service import ProfileService
from ..dependencies.auth import require_api_key

router = APIRouter(
    prefix="/profiles",
    tags=["profiles"],
    dependencies=[Depends(require_api_key)],
)

profile_service = ProfileService()

# Injected by main.py after app startup to avoid circular imports
increment_profiles_created: Callable[[], None] = lambda: None


def _owner(api_key: dict) -> Optional[str]:
    return api_key.get("id")


def _tier(api_key: dict) -> str:
    return api_key.get("tier", "free")


@router.post("/", response_model=Profile, status_code=201)
async def create_profile(
    profile_data: ProfileCreate,
    api_key: dict = Depends(require_api_key),
):
    """
    Create a new profile owned by the calling API key.

    Well-known ``profile_type`` values: personal, business, patient, provider,
    requesting_provider, attending_provider, billing_provider, pharmacy,
    facility, encounter, insured, agency, custom.
    """
    try:
        result = profile_service.create_profile(
            profile_data,
            tier=_tier(api_key),
            owner_id=_owner(api_key),
        )
        increment_profiles_created()
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create profile: {e}")


@router.get("/", response_model=List[Profile])
async def list_profiles(
    profile_type: Optional[str] = Query(None, description="Filter by profile_type"),
    api_key: dict = Depends(require_api_key),
):
    """List profiles visible to this API key (own + org-shared)."""
    return profile_service.list_profiles(
        owner_id=_owner(api_key),
        tier=_tier(api_key),
        profile_type=profile_type,
    )


@router.get("/{profile_id}", response_model=Profile)
async def get_profile(
    profile_id: str,
    api_key: dict = Depends(require_api_key),
):
    """Get profile by ID (must be owned or org-shared)."""
    profile = profile_service.get_profile(
        profile_id, owner_id=_owner(api_key), tier=_tier(api_key)
    )
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    return profile


@router.get(
    "/{profile_id}/fill-data",
    summary="Decrypted flat key/values for Guided Fill / template prefill",
)
async def get_profile_fill_data(
    profile_id: str,
    api_key: dict = Depends(require_api_key),
):
    """Return decrypted profile ``data`` (increments usage)."""
    try:
        return {
            "profile_id": profile_id,
            "data": profile_service.use_profile(
                profile_id, owner_id=_owner(api_key), tier=_tier(api_key)
            ),
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.patch("/{profile_id}", response_model=Profile)
async def update_profile(
    profile_id: str,
    update_data: ProfileUpdate,
    api_key: dict = Depends(require_api_key),
):
    """Update a profile you own."""
    try:
        return profile_service.update_profile(
            profile_id,
            update_data,
            owner_id=_owner(api_key),
            tier=_tier(api_key),
        )
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update profile: {e}")


@router.delete("/{profile_id}", status_code=204)
async def delete_profile(
    profile_id: str,
    api_key: dict = Depends(require_api_key),
):
    """Delete a profile you own."""
    try:
        if not profile_service.delete_profile(
            profile_id, owner_id=_owner(api_key), tier=_tier(api_key)
        ):
            raise HTTPException(status_code=404, detail="Profile not found")
    except ValueError:
        raise HTTPException(status_code=404, detail="Profile not found")
    return None


@router.post("/import", status_code=201, summary="Bulk-import profiles from CSV or Excel")
async def import_profiles(
    file: UploadFile = File(..., description="CSV or .xlsx file"),
    profile_type: Optional[str] = Form(
        None,
        description="Default profile_type for all rows (overridden by a 'profile_type' column in the file)",
    ),
    api_key: dict = Depends(require_api_key),
) -> Dict[str, Any]:
    """Import profiles owned by the calling API key."""
    fname = (file.filename or "").lower()
    content = await file.read()

    rows: List[Dict[str, str]] = []
    try:
        if fname.endswith(".csv"):
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader if any(v.strip() for v in r.values())]
        elif fname.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise HTTPException(400, "openpyxl not installed — use CSV format instead")
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            raw = list(ws.iter_rows(values_only=True))
            wb.close()
            if not raw:
                raise HTTPException(400, "Excel file is empty")
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(raw[0])]
            for row in raw[1:]:
                rec = {
                    headers[i]: str(v).strip() if v is not None else ""
                    for i, v in enumerate(row) if i < len(headers)
                }
                if any(v for v in rec.values()):
                    rows.append(rec)
        else:
            raise HTTPException(400, "File must be .csv or .xlsx")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Could not parse file: {exc}")

    if not rows:
        raise HTTPException(400, "File contains no data rows")
    if len(rows) > 500:
        raise HTTPException(400, "Maximum 500 rows per import")

    created, errors = [], []
    tier = _tier(api_key)
    owner_id = _owner(api_key)

    for i, row in enumerate(rows, start=2):
        name = (row.pop("name", "") or "").strip()
        if not name:
            errors.append({"row": i, "error": "Missing 'name' column"})
            continue
        shared_raw = (row.pop("shared", "") or "").strip().lower()
        org_id = (row.pop("org_id", "") or "").strip() or None
        ptype = (row.pop("profile_type", None) or profile_type or "personal").strip() or "personal"
        data = {k: v for k, v in row.items() if k and v}
        try:
            p = profile_service.create_profile(
                ProfileCreate(
                    name=name,
                    profile_type=ptype,
                    data=data,
                    shared=shared_raw in ("1", "true", "yes", "y"),
                    org_id=org_id,
                ),
                tier=tier,
                owner_id=owner_id,
            )
            increment_profiles_created()
            created.append(p.id)
        except ValueError as exc:
            errors.append({"row": i, "error": str(exc)})
        except Exception as exc:
            errors.append({"row": i, "error": f"Unexpected error: {exc}"})

    return {
        "created": len(created),
        "skipped": 0,
        "errors": errors,
        "profile_ids": created,
        "message": f"Imported {len(created)} profile{'s' if len(created) != 1 else ''}"
        + (f" — {len(errors)} row{'s' if len(errors) != 1 else ''} had errors" if errors else ""),
    }
